import json

# import requests
import boto3
import openpyxl
from pathlib import Path

old_sheet = "C:/Users/shiva/Downloads/server_consilidated_report.xlsx"
new_sheet = "C:/Users/shiva/Downloads/server_consilidated_report_today.xlsx"

out = 'E:/inventory.txt'


def report_v2():
    Output = open(out,'a')

    old_file = Path('Sheet', old_sheet)
    old_obj = openpyxl.load_workbook(old_file)
    old_sheet_1 = old_obj.active
    new_file = Path('Sheet', new_sheet)
    new_obj = openpyxl.load_workbook(new_file)
    new_sheet_1 = new_obj.active

    new_max_row = new_sheet_1.max_row
    old_max_row = old_sheet_1.max_row

    old_dict = {}
    new_dict = {}

    for i in range(2,old_max_row+1):
        if old_sheet_1[f"A{i}"].value in old_dict.keys():
            old_dict[old_sheet_1[f"A{i}"].value].append({"name": old_sheet_1[f"E{i}"].value, "state":old_sheet_1[f"D{i}"].value})
        else:
            old_dict[old_sheet_1[f"A{i}"].value] = [{"name": old_sheet_1[f"E{i}"].value, "state":old_sheet_1[f"D{i}"].value}]

    for i in range(2,new_max_row+1):
        if new_sheet_1[f"A{i}"].value in new_dict.keys():
            new_dict[new_sheet_1[f"A{i}"].value].append({"name": new_sheet_1[f"E{i}"].value, "state":new_sheet_1[f"D{i}"].value})
        else:
            new_dict[new_sheet_1[f"A{i}"].value] = [{"name": new_sheet_1[f"E{i}"].value, "state":new_sheet_1[f"D{i}"].value}]
    
    for key in new_dict.keys():
        log = []
        started = 0
        stopped = 0
        created = 0
        terminated = 0
        for item in new_dict[key]:
            found = False
            for old_item in old_dict[key]:
                if item["name"] == old_item["name"]:
                    if item["state"] == "running" and old_item["state"] == "stopped":
                        started += 1
                        log.append(f'{item["name"]} in {key} has been started')
                        #print(f'{item["name"]} in {key} has been started')
                        found = True
                    elif item["state"] == "stopped" and old_item["state"] == "running":
                        stopped += 1
                        log.append(f'{item["name"]} in {key} has been Stopped')
                        #print(f'{item["name"]} in {key} has been Started')
                        found = True
                    elif item["state"] == old_item["state"]:
                        found = True
                    else:
                        print(item["name"])
                    break
            if found == False:
                created += 1
                log.append(f'{item["name"]} in {key} has been Created')
                #print(f'{item["name"]} in {key} has been Created')
        for old_item in old_dict[key]:
            present = False
            for item in new_dict[key]:
                if old_item["name"] == item["name"]:
                    present = True
                    break
            if present == False:
                terminated += 1
                log.append(f'{old_item["name"]} in {key} has been Terminated')
        print(f'{key}: ',file = Output)
        print(f'\t Started : {started}',file = Output)
        print(f'\t Stopped : {stopped}',file = Output)
        print(f'\t created : {created}',file = Output)
        print(f'\t terminated : {terminated}',file = Output)
        for i in log:
            print(f'\t \t {i}',file = Output)

    print(f"Find the inventory in {out}")
    #main()
    

def lambda_handler(event, context):
    """Sample pure Lambda function

    Parameters
    ----------
    event: dict, required
        API Gateway Lambda Proxy Input Format

        Event doc: https://docs.aws.amazon.com/apigateway/latest/developerguide/set-up-lambda-proxy-integrations.html#api-gateway-simple-proxy-for-lambda-input-format

    context: object, required
        Lambda Context runtime methods and attributes

        Context doc: https://docs.aws.amazon.com/lambda/latest/dg/python-context-object.html

    Returns
    ------
    API Gateway Lambda Proxy Output Format: dict

        Return doc: https://docs.aws.amazon.com/apigateway/latest/developerguide/set-up-lambda-proxy-integrations.html
    """

    # try:
    #     ip = requests.get("http://checkip.amazonaws.com/")
    # except requests.RequestException as e:
    #     # Send some context about this error to Lambda Logs
    #     print(e)

    #     raise e

    report_v2()
